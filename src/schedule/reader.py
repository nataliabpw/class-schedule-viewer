from openpyxl import load_workbook
import pandas as pd
from .utils.profiler import timer

@timer
def load_spreadsheet_with_merged_cells(data_path, end_row):
    # Future improvement:
    # - try converting .xls to .xlsx using LibreOffice or other tool
    df = pd.read_excel(data_path, header=None, nrows=end_row)
    df = df.astype(object)

    wb = load_workbook(data_path)
    ws = wb.active # first sheet from file

    for merged_range in ws.merged_cells.ranges:

        merged_cell_value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        
        row_start = merged_range.min_row - 1
        row_end   = merged_range.max_row
        col_start = merged_range.min_col - 1
        col_end   = merged_range.max_col

        # df out-of-bounds protection
        if row_end <= df.shape[0] and col_end <= df.shape[1]:
            df.iloc[
                row_start:row_end,
                col_start:col_end
            ] = merged_cell_value
        elif str(merged_range)[:2]=='GB': 
            print(f"Warning: merged cell range {merged_range} is out of bounds for the dataframe and will be skipped.")
            print(f"rows {row_start}-{row_end}, columns {col_start}-{col_end}")
    return df