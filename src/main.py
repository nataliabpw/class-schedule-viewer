from datetime import datetime, date
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
import numpy as np

WEEKDAYS = ['PONIEDZIAŁEK', 'WTOREK', 'ŚRODA', 'CZWARTEK', 'PIĄTEK']

def main():
    project_root = Path(__file__).parent.parent
    data_dir = project_root / 'data'

    df = load_spreadsheet_with_merged_cells(data_dir)

    SCHEDULE_NAME = df.iloc[0,0]
    print(f"Schedule Name: {SCHEDULE_NAME}")

    date = get_date_from_user()
    weekday_id = date.weekday()  # Monday is 0 and Sunday is 6

    group_seminaria, group_cwiczenia, group_zajecia = get_group_names_from_user()
    
    if weekday_id > 4:
        print("Brak zajęć w weekend!")
        return
    
    print(f"Data: {date}\nDzień tygodnia: {WEEKDAYS[weekday_id]}")

    # Future-improvement:  Detect the row containing weekdays dynamically
    # urrently using hardcoded row index (1), which works only for the specific format
    weekday_row = 2

    weekday_start_column_id, weekday_end_column_id = find_columns_for_specific_weekday(df, weekday_id, weekday_row)
    # print(f"Columns: {weekday_start_column_id} - {weekday_end_column_id}")
    
    date_row = weekday_row + 2
    matching_date_columns = find_columns_with_matching_date(df, date, weekday_start_column_id, weekday_end_column_id, date_row)
    # print(matching_date_columns)

    class_name_row = weekday_row + 1
    start_row = weekday_row + 4
    end_row = df.shape[0]

    classes = build_group_schedule(class_name_row, start_row, end_row, matching_date_columns, df, group_seminaria, group_cwiczenia, group_zajecia)

    # Hardcoded, TO-DO: detect dynamically
    if weekday_id==0:
        time_column_id = weekday_start_column_id - 2
    else:
        time_column_id = weekday_start_column_id - 1 

    print_schedule_with_time(classes, df, start_row, time_column_id)


def load_spreadsheet_with_merged_cells(data_dir):
    # Future improvement:
    # - try converting .xls to .xlsx using LibreOffice or other tool
    converted_data_path = data_dir / 'converted_example_schedule.xlsx'
    df = pd.read_excel(converted_data_path, header=None, nrows=55)
    df = df.astype(object)

    wb = load_workbook(converted_data_path)
    ws = wb.active # first sheet from file

    for merged_range in ws.merged_cells.ranges:
        # min_col, min_row, max_col, max_row = merged_range.bounds

        merged_cell_value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        
        row_start = merged_range.min_row - 1
        row_end   = merged_range.max_row
        col_start = merged_range.min_col - 1
        col_end   = merged_range.max_col

        # zabezpieczenie przed wyjściem poza zakres df
        if row_end <= df.shape[0] and col_end <= df.shape[1]:
            df.iloc[
                row_start:row_end,
                col_start:col_end
            ] = merged_cell_value
    return df

def get_date_from_user():
    # print("Please enter a date: ")
    # day = input("Day: ")
    # month = input("Month: ")
    # year = input("Year: ")
    # return date(int(year), int(month), int(day))
    return date(2025, 11, 7) # Hardcoded for testing

def get_group_names_from_user():
    # print("Wprowadź grupy zajęć lub pozostaw puste, jeśli chcesz użyć domyślnych wartości:")
    # group_seminaria = input("Grupa na seminaria: ")
    # group_cwiczenia = input("Grupa na ćwiczenia: ")
    # grupa_zajecia = input("Grupa na zajęcia praktyczne: ")
    # if group_seminaria == "":
    #     group_seminaria = "1"
    # if group_cwiczenia == "":
    #     group_cwiczenia = "1a"
    # if group_zajecia == "":
    #     group_zajecia = "1a"
    
    # group_seminaria = 'grupa ' + group_seminaria.strip()
    # group_cwiczenia = ' ' + group_cwiczenia.strip()
    # group_zajecia = ' ' + group_zajecia.strip()

    # return group_seminaria, group_cwiczenia, group_zajecia
    return 'grupa 1', ' 1a', ' 1a' # Hardcoded for testing

def find_columns_for_specific_weekday(df, weekday_id, weekday_row):
    weekday_start_column_id = -1
    weekday_end_column_id = df.iloc[weekday_row].size-1

    for column_id, cell in enumerate(df.iloc[weekday_row]):
        if str(cell).strip() == WEEKDAYS[weekday_id] and not is_weekday_start_column_id_set(weekday_start_column_id):
            weekday_start_column_id = column_id
        elif is_weekday_start_column_id_set(weekday_start_column_id) and weekday_id==4:
            return weekday_start_column_id, weekday_end_column_id
        elif is_next_weekday_reached(weekday_start_column_id, cell, weekday_id):
            weekday_end_column_id = column_id - 3
            return weekday_start_column_id, weekday_end_column_id
    return weekday_start_column_id, weekday_end_column_id

def find_columns_with_matching_date(df, date, weekday_start_column_id, weekday_end_column_id, date_row):
    matching_date_columns = []
    for column_id, cell in enumerate(df.iloc[date_row, weekday_start_column_id:weekday_end_column_id+1]):
        current_column_id = column_id+weekday_start_column_id

        if isinstance(cell, datetime):
            if cell.date() == date:
                matching_date_columns.append(current_column_id)
            continue
        cell = str(cell).strip()

        DASH = '-'
        if DASH in cell:
            dash_index = cell.index(DASH)
            start_date = cell[0:dash_index].strip()
            end_date = cell[dash_index+1:].strip()

            # zamiana str na datę
            start_date = format_date(start_date, date)
            end_date = format_date(end_date, date)
            # obsługa stycznia
            if start_date>end_date:
                end_date = end_date.replace(year=end_date.year + 1)

            # (sprawdzenie czy zakres pasuje do daty)
            if start_date <= date <= end_date:
                # TO-DO: Obsługa komentarza "bez dd.mm"
                if is_date_an_exception(cell, date):
                    print(cell)
                    continue
                # (zapamiętanie kolumny)
                matching_date_columns.append(current_column_id)
        elif cell=='cały semestr':
            matching_date_columns.append(current_column_id )
        elif cell != 'nan':
            date_from_cell = format_date(cell, date)
            if date_from_cell==date:
                matching_date_columns.append(current_column_id )
    return matching_date_columns

def is_next_weekday_reached(weekday_start_column_id, cell, weekday_id):
    if is_weekday_start_column_id_set(weekday_start_column_id) and str(cell).strip() == WEEKDAYS[weekday_id+1]:
        return True
    return False

def is_weekday_start_column_id_set(weekday_start_column_id):
    if weekday_start_column_id==-1:
        return False
    return True

def is_date_an_exception(cell, date):
    cell = cell.lower()
    if 'bez' not in cell:
        return False
    BEZ_LENGTH = 3
    bez_end_id = cell.index('bez') + BEZ_LENGTH
    exceptions = cell[bez_end_id:]
    date_to_check = date.strftime("%d.%m")
    if date_to_check in exceptions:
        return True
    return False

def format_date(date_to_format, date_from_user):
    if date_to_format[-1]==')':
        id = date_to_format.index('(')
        date_to_format = date_to_format[:id].strip()
    if date_to_format[-1]!='.':
        date_to_format += '.'
    if len(date_to_format)<10:
        if is_date_from_user_from_next_year(date_to_format, date_from_user):
            date_to_format += str(date_from_user.year-1)
        else:
            date_to_format += str(date_from_user.year)
    return datetime.strptime(date_to_format, "%d.%m.%Y").date()

def is_date_from_user_from_next_year(date_to_format, date_from_user):
    if int(date_to_format[-3:-1]) - date_from_user.month > 7:
        return True
    return False

def build_group_schedule(class_name_row, start_row, end_row, matching_date_columns, df, group_seminaria, group_cwiczenia, group_zajecia):
    classes = [None] * (end_row - start_row)

    for column_id in matching_date_columns:
        if 'ćwiczenia' in df.iloc[class_name_row, column_id].lower():
            group = group_cwiczenia
        elif 'zajęcia praktyczne' in df.iloc[class_name_row, column_id].lower():
            group = group_zajecia
        else:
            group = group_seminaria
        for row_id in range(start_row, end_row):
            cell = str(df.iloc[row_id, column_id]).strip()
            if group in cell:
                if cell.index(group)+len(group) < len(cell):
                    if cell[cell.index(group)+len(group)] in '0123456789':
                        continue
                classes[row_id-start_row] = df.iloc[class_name_row, column_id].strip()+" - "+cell
    return classes

def print_schedule_with_time(classes, df, start_row, time_column_id):
    last_class = None
    for row_id, curr_class in enumerate(classes):
        # wypisz koniec zajęć
        if last_class is not None and curr_class!=last_class:
            end_time = df.iloc[row_id+start_row-1, time_column_id]
            end_time = end_time[end_time.index('-')+1:].strip()
            print(f"Koniec: {end_time}")
        # wypisz początek zajęć
        if curr_class is not None and curr_class!=last_class:
            print(f"Zajęcia: {curr_class}")
            start_time = df.iloc[row_id+start_row, time_column_id]
            start_time = start_time[:start_time.index('-')].strip()
            print(f"Początek: {start_time}")
        last_class = curr_class

if __name__ == "__main__":
    main()